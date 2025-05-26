from tkinter import *
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

a = Tk()

def reset():
    input_name.set("")
    input_age.set("")
    input_contact.set("")
    input_add.set("")
    gender_entry.set("Select")
    result_label.config(text="")  # Clear the result

def submit():
    wb = Workbook() if not pathlib.Path("data.xlsx").exists() else openpyxl.load_workbook("data.xlsx")
    ws = wb.active
    if ws.max_row == 1:
        ws.append(["Name", "Age", "Contact", "Address", "Gender"])
    ws.append([input_name.get(), input_age.get(), input_contact.get(), input_add.get(), gender_entry.get()])
    wb.save("data.xlsx")

a.title("Data Entry Form")
a.geometry("300x350")

Label(a, text="Name:", font=("arial", 10, "bold")).grid(row=0, column=0, padx=10, pady=5)
Label(a, text="Age:", font=("arial", 10, "bold")).grid(row=1, column=0, padx=10, pady=5)
Label(a, text="Contact:", font=("arial", 10, "bold")).grid(row=2, column=0, padx=10, pady=5)
Label(a, text="Address:", font=("arial", 10, "bold")).grid(row=3, column=0, padx=10, pady=5)
Label(a, text="Gender:", font=("arial", 10, "bold")).grid(row=4, column=0, padx=10, pady=5)

input_name = StringVar()
input_age = StringVar()
input_contact = StringVar()
input_add = StringVar()

Entry(a, font=("arial", 10, "bold"), textvariable=input_name, bg="#AE7EA0", width=20).grid(row=0, column=1)
Entry(a, font=("arial", 10, "bold"), textvariable=input_age, bg="#AE7EA0", width=20).grid(row=1, column=1)
Entry(a, font=("arial", 10, "bold"), textvariable=input_contact, bg="#AE7EA0", width=20).grid(row=2, column=1)
Entry(a, font=("arial", 10, "bold"), textvariable=input_add, bg="#AE7EA0", width=20).grid(row=3, column=1)

gender_entry = Combobox(a, values=['Male', 'Female'], font="arial 10", state='readonly', width=17)
gender_entry.grid(row=4, column=1, padx=2, pady=3)
gender_entry.set("Select")

s_btn = Button(a, text="Submit", command=submit)
s_btn.grid(row=5, column=0, padx=5, pady=10)

r_btn = Button(a, text="Reset", command=reset)
r_btn.grid(row=5, column=1, padx=5, pady=10)

e_btn = Button(a, text="Exit", command=lambda: a.destroy())
e_btn.grid(row=6, column=0, columnspan=2, pady=10)

# Label to display result
result_label = Label(a, text="", font=("arial", 9), justify=LEFT)
result_label.grid(row=7, column=0, columnspan=2, pady=10)

a.mainloop()